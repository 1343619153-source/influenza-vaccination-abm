import math
import os
import random
import re
from datetime import datetime
import matplotlib.pyplot as plt
import pandas as pd
import warnings
import numpy as np
import multiprocessing as mp
import copy
from dataclasses import dataclass
from typing import Any, Dict, List, Optional, Sequence, Tuple, Union

PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))

try:
    from scipy import stats

    HAS_SCIPY = True
except ImportError:
    HAS_SCIPY = False
    print("Warning: scipy is not installed; approximate t-distribution values will be used for confidence intervals")

warnings.filterwarnings('ignore')

plt.rcParams['font.sans-serif'] = ['DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

SIMULATION_MONTHS = 8

SIMULATION_MONTH_LABELS = ['Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']
DAYS_PER_MONTH = 30
STEPS_PER_MONTH = DAYS_PER_MONTH * 2
HALF_MONTH_STEPS = STEPS_PER_MONTH // 2
SIMULATION_DAYS = SIMULATION_MONTHS * DAYS_PER_MONTH
DEFAULT_SIMULATION_STEPS = SIMULATION_MONTHS * STEPS_PER_MONTH

Number = Union[int, float]


def _ili_lambda_as_float(x: Number, name: str) -> float:
    v = float(x)
    if not math.isfinite(v):
        raise ValueError(f"{name} must be a finite real number")
    return v


@dataclass(frozen=True)
class AgeRates:
    lambda_child: float
    lambda_adult: float
    lambda_elderly: float


def _population_weights_for_ili_lambda(
    pi_child: Number,
    pi_adult: Number,
    pi_elderly: Number,
    r_child: Number,
    r_adult: Number,
    r_elderly: Number,
) -> Tuple[float, float, float, float]:
    pc, pa, pe = _ili_lambda_as_float(pi_child, "pi_child"), _ili_lambda_as_float(pi_adult, "pi_adult"), _ili_lambda_as_float(
        pi_elderly, "pi_elderly"
    )
    if pc < 0 or pa < 0 or pe < 0:
        raise ValueError("Population proportions must be nonnegative")
    s = pc + pa + pe
    if abs(s - 1.0) > 1e-6:
        raise ValueError(f"pi_c + pi_a + pi_e must equal 1; got {s}")

    rc, ra, re = _ili_lambda_as_float(r_child, "r_child"), _ili_lambda_as_float(r_adult, "r_adult"), _ili_lambda_as_float(
        r_elderly, "r_elderly"
    )
    if rc <= 0 or ra <= 0 or re <= 0:
        raise ValueError("Published attack rates r must be positive")

    wc, wa, we = rc * pc, ra * pa, re * pe
    w_sum = wc + wa + we
    if w_sum <= 0:
        raise ValueError("W = sum(r_a * pi_a) must be positive")
    return w_sum, wc, wa, we


def infection_rates_from_totals(
    ili_total: Number,
    ili_positivity: Number,
    pi_child: Number,
    pi_adult: Number,
    pi_elderly: Number,
    h_child: Number,
    h_adult: Number,
    h_elderly: Number,
    *,
    calibration_c: Number = 1.0,
    r_child: Number,
    r_adult: Number,
    r_elderly: Number,
) -> AgeRates:
    ili = _ili_lambda_as_float(ili_total, "ili_total")
    pos = _ili_lambda_as_float(ili_positivity, "ili_positivity")
    if ili < 0 or pos < 0 or pos > 1:
        raise ValueError("ILI_tot must be nonnegative and p_pos must be in [0, 1]")

    hc = _ili_lambda_as_float(h_child, "h_child")
    ha = _ili_lambda_as_float(h_adult, "h_adult")
    he = _ili_lambda_as_float(h_elderly, "h_elderly")
    for name, h in [("h_child", hc), ("h_adult", ha), ("h_elderly", he)]:
        if h <= 0 or h > 1:
            raise ValueError(f"{name} must be in (0, 1]")

    c_cal = _ili_lambda_as_float(calibration_c, "calibration_c")
    if c_cal <= 0:
        raise ValueError("calibration_c must be positive")

    w_sum, _, _, _ = _population_weights_for_ili_lambda(pi_child, pi_adult, pi_elderly, r_child, r_adult, r_elderly)

    rc = _ili_lambda_as_float(r_child, "r_child")
    ra = _ili_lambda_as_float(r_adult, "r_adult")
    re = _ili_lambda_as_float(r_elderly, "r_elderly")

    base = c_cal * ili * pos / w_sum
    lam_c = base * rc / hc
    lam_a = base * ra / ha
    lam_e = base * re / he

    return AgeRates(lambda_child=lam_c, lambda_adult=lam_a, lambda_elderly=lam_e)


def create_node_network(family_dist=None, rng=None):
    if rng is None:
        rng = random

    children_count = POPULATION_CONFIG['children_count']
    elderly_count = POPULATION_CONFIG['elderly_count']

    if family_dist is not None:
        parents_count = family_dist.get('parents_used', 0)
        singles_count = family_dist.get('singles_used', 0)
    else:
        if 'parents_count' in POPULATION_CONFIG and 'singles_count' in POPULATION_CONFIG:
            parents_count = POPULATION_CONFIG['parents_count']
            singles_count = POPULATION_CONFIG['singles_count']
        else:
            adults_count = POPULATION_CONFIG.get('adults_count', 0)

            parents_count = int(adults_count * 0.7)
            singles_count = adults_count - parents_count

    total_nodes = children_count + parents_count + singles_count + elderly_count

    nodes = {}
    node_id = 0

    child_indices = []
    for i in range(children_count):
        nodes[node_id] = {
            'type': 'child',
            'name': f'Child {i + 1}',
            'node_id': node_id,
            'parents': [],
            'family_id': None,
            'health_status': None
        }
        child_indices.append(node_id)
        node_id += 1

    parent_indices = []
    parent_pairs = []
    for i in range(parents_count):
        partner_num = i // 2

        nodes[node_id] = {
            'type': 'parent',
            'name': f'Father {partner_num + 1}' if i % 2 == 0 else f'Mother {partner_num + 1}',
            'node_id': node_id,
            'family_id': partner_num,
            'health_status': None
        }
        parent_indices.append(node_id)

        if i % 2 == 1:
            parent_pairs.append([node_id - 1, node_id])

        node_id += 1

    family_id = 0
    for i, child_idx in enumerate(child_indices):
        if i < len(parent_pairs):
            nodes[child_idx]['family_id'] = family_id
            nodes[child_idx]['parents'] = [f'Father {family_id + 1}', f'Mother {family_id + 1}']
            family_id += 1
        else:
            max_attempts = len(parent_pairs) * 2
            attempts = 0
            assigned = False
            while not assigned and attempts < max_attempts:
                assigned_family = rng.randint(0, len(parent_pairs) - 1)

                family_children = [c for c in child_indices if nodes[c]['family_id'] == assigned_family]
                if len(family_children) < 2:
                    nodes[child_idx]['family_id'] = assigned_family
                    nodes[child_idx]['parents'] = [f'Father {assigned_family + 1}', f'Mother {assigned_family + 1}']
                    assigned = True
                attempts += 1

            if not assigned:
                assigned_family = rng.randint(0, len(parent_pairs) - 1)
                nodes[child_idx]['family_id'] = assigned_family
                nodes[child_idx]['parents'] = [f'Father {assigned_family + 1}', f'Mother {assigned_family + 1}']

    single_indices = []
    for i in range(singles_count):
        nodes[node_id] = {
            'type': 'single',
            'name': f'Single adult {i + 1}',
            'node_id': node_id,
            'family_id': None,
            'health_status': None
        }
        single_indices.append(node_id)
        node_id += 1

    elderly_indices = []
    for i in range(elderly_count):
        nodes[node_id] = {
            'type': 'elderly',
            'name': f'Older adult {i + 1}',
            'node_id': node_id,
            'family_id': None,
            'health_status': None
        }
        elderly_indices.append(node_id)
        node_id += 1

    return nodes, child_indices, parent_indices, single_indices, elderly_indices


def add_school_connections(nodes, child_indices, rng=None):
    if rng is None:
        rng = random

    total_children = len(child_indices)
    categories = NETWORK_CONFIG['school_categories']

    total_ratio = sum(cat.get('student_ratio', 0.0) for cat in categories)
    if total_ratio <= 0:
        raise ValueError("The sum of school-category student_ratio values must be positive")
    if abs(total_ratio - 1.0) > 0.01:
        print(f"Warning: school-category student ratios sum to {total_ratio:.3f}, not 1.0; normalizing")
        categories = [dict(cat, student_ratio=cat.get('student_ratio', 0.0) / total_ratio) for cat in categories]

    school_assignments_list = []
    school_index = 0
    for cat in categories:
        category_name = cat.get('name', 'Unknown category')
        school_count = int(cat.get('count', 0))
        if school_count <= 0:
            raise ValueError(f"School category {category_name} must have a positive count")
        category_students = int(total_children * float(cat.get('student_ratio', 0.0)))
        students_per_school = category_students // school_count
        remainder = category_students % school_count
        for i in range(school_count):
            school_students = students_per_school + (1 if i < remainder else 0)
            school_assignments_list.append({
                'name': f"{category_name}{i + 1}",
                'count': school_students,
                'index': school_index,
                'category': category_name,
                'connections': cat['connections'],
            })
            school_index += 1

    total_assigned = sum(s['count'] for s in school_assignments_list)
    if total_assigned < total_children:
        school_assignments_list[0]['count'] += total_children - total_assigned
    elif total_assigned > total_children:
        diff = total_assigned - total_children
        for school in reversed(school_assignments_list):
            take = min(diff, max(0, school['count']))
            school['count'] -= take
            diff -= take
            if diff <= 0:
                break

    school_assignments = []
    for school_info in school_assignments_list:
        school_assignments.extend([school_info] * school_info['count'])
    rng.shuffle(school_assignments)

    for i, child_id in enumerate(child_indices):
        school_info = school_assignments[i] if i < len(school_assignments) else rng.choice(school_assignments_list)
        nodes[child_id]['school'] = school_info['name']
        nodes[child_id]['school_number'] = school_info['index'] + 1

    print("\n=== School assignment by size category ===")
    for cat in categories:
        category_name = cat.get('name', 'unknown')
        category_schools = [s for s in school_assignments_list if s['category'] == category_name]
        total_students = sum(s['count'] for s in category_schools)
        ratio = total_students / total_children * 100 if total_children else 0.0
        connections = cat['connections']
        print(f"{category_name}: {len(category_schools)} schools, {total_students} students ({ratio:.2f}%), connections: {connections[0]}-{connections[1]}")
        for school in category_schools:
            print(f"  - {school['name']}: {school['count']} students")

    school_children = {}
    school_category_map = {school['name']: school for school in school_assignments_list}
    for node_id, node_info in nodes.items():
        if node_info['type'] == 'child':
            school_children.setdefault(node_info['school'], []).append(node_id)

    for school, children in school_children.items():
        min_conn, max_conn = school_category_map[school]['connections']
        for child_id in children:
            num_connections = rng.randint(min_conn, max_conn)
            other_children = [c for c in children if c != child_id]
            connections = rng.sample(other_children, num_connections) if len(other_children) >= num_connections else other_children
            nodes[child_id]['school_connections'] = connections

    return nodes


def add_company_connections(nodes, rng=None):
    if rng is None:
        rng = random

    adults = [node_id for node_id, node_info in nodes.items()
              if node_info['type'] in ['parent', 'single']]

    num_groups = NETWORK_CONFIG['num_work_groups']
    adults_per_group = len(adults) // num_groups

    group_assignments = []
    for group_num in range(1, num_groups + 1):
        group_assignments.extend([group_num] * adults_per_group)

    rng.shuffle(group_assignments)

    for i, adult_id in enumerate(adults):
        group_num = group_assignments[i] if i < len(group_assignments) else rng.randint(1, num_groups)
        nodes[adult_id]['group'] = f'Group {group_num}'
        nodes[adult_id]['group_number'] = group_num

    company_adults = {}
    for node_id, node_info in nodes.items():
        if node_info['type'] in ['parent', 'single'] and 'group' in node_info:
            company = node_info['group']
            if company not in company_adults:
                company_adults[company] = []
            company_adults[company].append(node_id)

    for company, adults in company_adults.items():
        for adult_id in adults:
            min_conn, max_conn = NETWORK_CONFIG['adults_connections']
            num_connections = rng.randint(min_conn, max_conn)

            other_adults = [a for a in adults if a != adult_id]
            if len(other_adults) >= num_connections:
                connections = rng.sample(other_adults, num_connections)
            else:
                connections = other_adults

            nodes[adult_id]['company_connections'] = connections
    return nodes


def add_elderly_connections(nodes, rng=None):
    if rng is None:
        rng = random

    elderly_nodes = [node_id for node_id, node_info in nodes.items() if node_info['type'] == 'elderly']

    for elderly_id in elderly_nodes:
        min_conn, max_conn = NETWORK_CONFIG['elderly_connections']
        num_connections = rng.randint(min_conn, max_conn)

        other_elderly = [e for e in elderly_nodes if e != elderly_id]
        if len(other_elderly) >= num_connections:
            connections = rng.sample(other_elderly, num_connections)
        else:
            connections = other_elderly

        nodes[elderly_id]['elderly_connections'] = connections
    return nodes


def assign_health_status(nodes, rng=None):
    if rng is None:
        rng = random

    total_nodes = len(nodes)
    for node_id in nodes.keys():
        nodes[node_id]['health_status'] = 'susceptible'
        nodes[node_id]['ever_vaccinated'] = False
        nodes[node_id]['post_vaccine_susceptible'] = False
        nodes[node_id].pop('vacc_antibody_step', None)

    summary = {
        'total': total_nodes,
        'susceptible': total_nodes,
        'infected': 0,
        'vaccinated': 0,
        'recovered': 0,
    }

    return nodes, summary


def _node_population_group(node_type: str) -> str:
    if node_type == 'child':
        return 'child'
    if node_type == 'elderly':
        return 'elderly'
    return 'adult'


def assign_prior_vaccination_history(nodes, rng=None) -> Dict[str, int]:
    cfg = PRIOR_VACCINATION_CONFIG or {}
    counts = {'child': 0, 'adult': 0, 'elderly': 0, 'total': 0}
    if not cfg.get('enabled', True):
        for node_id in nodes:
            nodes[node_id]['prior_season_vaccinated'] = False
        return counts

    if rng is None:
        rng = random

    rate_by_group = {
        'child': float(cfg.get('coverage_child', 0.0)),
        'adult': float(cfg.get('coverage_adult', 0.0)),
        'elderly': float(cfg.get('coverage_elderly', 0.0)),
    }
    for node_id, node in nodes.items():
        grp = _node_population_group(node.get('type', ''))
        p_prior = max(0.0, min(1.0, rate_by_group.get(grp, 0.0)))
        had_prior = rng.random() < p_prior
        node['prior_season_vaccinated'] = bool(had_prior)
        if had_prior:
            counts[grp] += 1
            counts['total'] += 1
    return counts


def print_prior_vaccination_summary(nodes) -> None:
    cfg = PRIOR_VACCINATION_CONFIG or {}
    if not cfg.get('enabled', True):
        print("\nPrior-season vaccination history is disabled (PRIOR_VACCINATION_CONFIG['enabled']=False)")
        return
    totals = {'child': 0, 'adult': 0, 'elderly': 0}
    prior = {'child': 0, 'adult': 0, 'elderly': 0}
    for node in nodes.values():
        grp = _node_population_group(node.get('type', ''))
        totals[grp] += 1
        if node.get('prior_season_vaccinated'):
            prior[grp] += 1
    print("\n=== Prior-Season Vaccination History (prior_season_vaccinated) ===")
    for grp, label in [('child', 'Children'), ('adult', 'Adults'), ('elderly', 'Older adults')]:
        n = totals[grp]
        k = prior[grp]
        pct = k / n * 100 if n else 0.0
        target = float(cfg.get(f'coverage_{grp}', 0.0)) * 100
        print(f"  {label}: {k}/{n} ({pct:.1f}%); configured target: approximately {target:.1f}%")


def print_initial_health_summary(summary):
    total = summary['total']
    print("\n=== Initial Health-State Distribution ===")
    for key, label in [('susceptible', 'Susceptible'), ('infected', 'Infected'), ('vaccinated', 'Vaccinated'), ('recovered', 'Recovered')]:
        count = summary[key]
        percentage = count / total * 100 if total else 0
        print(f"{label}: {count} people ({percentage:.2f}%)")


def get_daytime_connections(node_id, nodes):
    node_info = nodes[node_id]
    connections = []

    if node_info['type'] == 'child':
        if 'school_connections' in node_info:
            connections = node_info['school_connections']
    elif node_info['type'] in ['parent', 'single']:
        if 'company_connections' in node_info:
            connections = node_info['company_connections']
    elif node_info['type'] == 'elderly':
        if 'elderly_connections' in node_info:
            connections = node_info['elderly_connections']

    return connections


def get_nighttime_connections(node_id, nodes):
    node_info = nodes[node_id]
    connections = []

    family_id = node_info.get('family_id')

    if family_id is not None:
        for other_id, other_info in nodes.items():
            if other_id != node_id and other_info.get('family_id') == family_id:
                connections.append(other_id)

    return connections


def calculate_neighbor_infection_probability(node_id, nodes, month_params=None, step=None):
    if step is None:
        if nodes[node_id]['health_status'] != 'susceptible':
            return 0.0
    elif not _can_receive_infection(nodes, node_id, step):
        return 0.0

    connections = get_daytime_connections(node_id, nodes)
    if not connections:
        return 0.0

    x = sum(
        1
        for cid in connections
        if cid in nodes and nodes[cid]['health_status'] == 'infected'
    )
    if x == 0:
        return 0.0

    mp = month_params or {}
    p_n = float(mp.get('p_single_infection', INFECTION_CONFIG.get('p_single_infection', 0.05)))
    p_n = max(0.0, min(1.0, p_n))
    return 1.0 - (1.0 - p_n) ** x


def autonomous_infection_rate_for_node(node_id, nodes, month_params=None, lam_rates=None):
    if lam_rates is None:
        return 0.0
    node_type = nodes[node_id].get('type')
    if node_type == 'child':
        v = lam_rates.lambda_child
    elif node_type in ('parent', 'single'):
        v = lam_rates.lambda_adult
    elif node_type == 'elderly':
        v = lam_rates.lambda_elderly
    else:
        v = 0.0
    return float(max(0.0, min(1.0, v)))


def _can_receive_infection(nodes, node_id, step: int) -> bool:
    st = nodes[node_id].get('health_status')
    if st == 'susceptible':
        return True
    if st == 'vaccinated':
        return nodes[node_id].get('vacc_antibody_step') is not None
    return False


def _vaccine_infection_multiplier(nodes, node_id, step: int) -> float:
    if nodes[node_id].get('health_status') != 'vaccinated':
        return 1.0
    ab = nodes[node_id].get('vacc_antibody_step')
    if ab is None or step < ab:
        return 1.0
    nt = nodes[node_id].get('type')
    cfg = VACCINE_CONFIG
    if nt == 'child':
        ve = cfg['ve_child']
    elif nt in ('parent', 'single'):
        ve = cfg['ve_adult']
    elif nt == 'elderly':
        ve = cfg['ve_elderly']
    else:
        ve = 0.0
    return max(0.0, 1.0 - float(ve))


def calculate_payoffs(nodes, is_daytime, month_params=None, step: int = 1):
    all_nodes = list(nodes.keys())

    for node_id in all_nodes:
        if is_daytime:
            connections = get_daytime_connections(node_id, nodes)
        else:
            connections = get_nighttime_connections(node_id, nodes)

        if not connections:
            nodes[node_id]['group_payoff'] = 0
            nodes[node_id]['individual_payoff'] = 0
            continue

        N = len(connections)
        N_I = sum(1 for conn_id in connections if nodes[conn_id]['health_status'] == 'infected')
        N_V = sum(1 for conn_id in connections if nodes[conn_id]['health_status'] == 'vaccinated')
        N_R = 0

        infection_risk_connections = [
            conn_id for conn_id in connections if _can_receive_infection(nodes, conn_id, step)
        ]
        sum_P_inf = 0
        for conn_id in infection_risk_connections:
            sum_P_inf += calculate_neighbor_infection_probability(conn_id, nodes, month_params, step)

        C_I = COST_CONFIG['infection_cost']
        C_V = COST_CONFIG['vaccination_cost']
        C_R = COST_CONFIG['recovery_cost']

        group_payoff = -((N_I / N * C_I) + (N_V / N * C_V) + (C_I * sum_P_inf / N) + (N_R / N * C_R))

        if nodes[node_id]['health_status'] == 'vaccinated':
            individual_payoff = group_payoff - C_V
        elif nodes[node_id]['health_status'] == 'infected':
            individual_payoff = group_payoff - C_I
        else:
            individual_payoff = group_payoff

        nodes[node_id]['group_payoff'] = group_payoff
        nodes[node_id]['individual_payoff'] = individual_payoff


def _sigmoid(x):
    x = max(-60.0, min(60.0, x))
    return 1.0 / (1.0 + math.exp(-x))


def _parse_percent_value(value):
    if pd.isna(value):
        return np.nan
    if isinstance(value, (int, float, np.number)):
        value = float(value)
        return value / 100.0 if value > 1 else value
    text = str(value).strip()
    if text.endswith('%'):
        text = text[:-1]
        return float(text) / 100.0
    value = float(text)
    return value / 100.0 if value > 1 else value


def day_index_from_step(step: int) -> int:
    return (step - 1) // 2 + 1


def week_index_from_step(step: int) -> int:
    d = day_index_from_step(step)
    return (d - 1) // 7 + 1

SIMULATION_EPI_WEEK_SEQUENCE = list(range(31, 53)) + list(range(1, 14))


def csv_epi_week_from_simulation_week(sim_week: int) -> int:
    if sim_week < 1:
        sim_week = 1
    seq = SIMULATION_EPI_WEEK_SEQUENCE
    return seq[(sim_week - 1) % len(seq)]


def _read_weekly_ili_csv(filepath: str) -> pd.DataFrame:
    try:
        return pd.read_csv(filepath, encoding='utf-8-sig')
    except UnicodeDecodeError:
        try:
            return pd.read_csv(filepath, encoding='gbk')
        except UnicodeDecodeError:
            return pd.read_csv(filepath, encoding='gb18030')


def _lookup_lambda_rates_from_cache(step: int, cache) -> Optional[AgeRates]:
    if not cache:
        return None
    sim_week = week_index_from_step(step)
    epi_w = csv_epi_week_from_simulation_week(sim_week)
    row = cache.get(epi_w)
    if not row:
        return None

    cfg = ILI_LAMBDA_FORMULA_CONFIG
    return infection_rates_from_totals(
        ili_total=row['ili_tot'],
        ili_positivity=row['p_pos'],
        pi_child=cfg['pi_child'],
        pi_adult=cfg['pi_adult'],
        pi_elderly=cfg['pi_elderly'],
        h_child=cfg['h_child'],
        h_adult=cfg['h_adult'],
        h_elderly=cfg['h_elderly'],
        calibration_c=cfg['calibration_c'],
        r_child=cfg['attack_child'],
        r_adult=cfg['attack_adult'],
        r_elderly=cfg['attack_elderly'],
    )


def _ili_csv_fail(reason: str):
    global ILI_WEEKLY_CSV_LOAD_FAILED, VACCINATION_ILI_WEEKLY_LOOKUP_CACHE, ILI_AUTONOMOUS_WEEK_CACHE
    ILI_WEEKLY_CSV_LOAD_FAILED = True
    VACCINATION_ILI_WEEKLY_LOOKUP_CACHE = {}
    ILI_AUTONOMOUS_WEEK_CACHE = {}
    print(f"Failed to read weekly ILI CSV: {reason}")


def load_weekly_ili_caches():
    global VACCINATION_ILI_WEEKLY_LOOKUP_CACHE, ILI_AUTONOMOUS_WEEK_CACHE, ILI_WEEKLY_CSV_LOAD_FAILED

    ILI_WEEKLY_CSV_LOAD_FAILED = False
    VACCINATION_ILI_WEEKLY_LOOKUP_CACHE = {}
    ILI_AUTONOMOUS_WEEK_CACHE = {}

    path = ILI_CONFIG.get('weekly_csv') or ''
    if not path or not os.path.isfile(path):
        _ili_csv_fail("weekly_csv is not configured or the file does not exist")
        return

    try:
        df = _read_weekly_ili_csv(path)
    except Exception as e:
        _ili_csv_fail(str(e))
        return

    if df.shape[0] == 0 or df.shape[1] < 2:
        _ili_csv_fail("the file is empty or has too few columns")
        return

    default_pos = float(ILI_CONFIG.get('default_weekly_positivity', 0.25))

    if df.shape[1] >= 9:
        rows_ordered = []
        for _, row in df.iterrows():
            try:
                wk = int(float(row.iloc[0]))
            except Exception:
                continue
            p25 = _parse_percent_value(row.iloc[7])
            h24 = _parse_percent_value(row.iloc[5])
            pos25 = _parse_percent_value(row.iloc[8])
            if pd.isna(p25):
                continue
            if pd.isna(h24):
                h24 = p25
            if pd.isna(pos25):
                pos25 = default_pos

            rows_ordered.append(
                {
                    'wk': wk,
                    'pred25': float(p25),
                    'hist24': float(h24),
                    'pos25': float(pos25),
                }
            )

        if not rows_ordered:
            _ili_csv_fail("the wide-format table contains no valid data rows")
            return

        lookup = {}
        for i, rec in enumerate(rows_ordered):
            wk = rec['wk']
            p25 = rec['pred25']
            h24 = rec['hist24']
            pos25 = rec['pos25']

            if i > 0:
                p_prev = rows_ordered[i - 1]['pred25']
                ili_trend = (p25 - p_prev) / (p_prev + 1e-6)
            else:
                ili_trend = 0.0

            ili_relative = max(0.0, (p25 - h24) / (h24 + 1e-6))

            lookup[wk] = {'ili_trend': float(ili_trend), 'ili_relative': float(ili_relative)}
            ILI_AUTONOMOUS_WEEK_CACHE[wk] = {'ili_tot': p25, 'p_pos': pos25}

        VACCINATION_ILI_WEEKLY_LOOKUP_CACHE = lookup
        ILI_WEEKLY_CSV_LOAD_FAILED = False
        return

    weeks = []
    preds = []
    hists = []
    for _, row in df.iterrows():
        try:
            wk = int(float(row.iloc[0]))
        except Exception:
            continue
        pv = _parse_percent_value(row.iloc[1])
        if pd.isna(pv):
            continue
        hv = _parse_percent_value(row.iloc[2]) if df.shape[1] > 2 else np.nan
        if pd.isna(hv):
            hv = pv

        if df.shape[1] >= 4:
            pos_v = _parse_percent_value(row.iloc[3])
            if pd.isna(pos_v):
                pos_v = default_pos
        else:
            pos_v = default_pos

        ILI_AUTONOMOUS_WEEK_CACHE[wk] = {'ili_tot': float(pv), 'p_pos': float(pos_v)}

        weeks.append(wk)
        preds.append(float(pv))
        hists.append(float(hv))

    if not weeks:
        _ili_csv_fail("the narrow-format table contains no valid weekly data rows")
        return

    max_w = max(weeks)
    pred_by_w = [np.nan] * (max_w + 1)
    hist_by_w = [np.nan] * (max_w + 1)
    for wk, p, h in zip(weeks, preds, hists):
        if 1 <= wk <= max_w:
            pred_by_w[wk] = p
            hist_by_w[wk] = h

    lookup = {}
    for w in range(1, max_w + 1):
        pw = pred_by_w[w]
        if pw is None or (isinstance(pw, float) and np.isnan(pw)):
            continue
        if w >= 2:
            pw_prev = pred_by_w[w - 1]
            if pw_prev is None or (isinstance(pw_prev, float) and np.isnan(pw_prev)):
                ili_trend = 0.0
            else:
                ili_trend = (pw - pw_prev) / (pw_prev + 1e-6)
        else:
            ili_trend = 0.0

        hw = hist_by_w[w]
        if hw is None or (isinstance(hw, float) and np.isnan(hw)):
            ili_relative = 0.0
        else:
            ili_relative = max(0.0, (pw - hw) / (hw + 1e-6))

        lookup[w] = {'ili_trend': float(ili_trend), 'ili_relative': float(ili_relative)}

    VACCINATION_ILI_WEEKLY_LOOKUP_CACHE = lookup
    ILI_WEEKLY_CSV_LOAD_FAILED = False


def load_weekly_ili_feature_lookup():
    load_weekly_ili_caches()
    return VACCINATION_ILI_WEEKLY_LOOKUP_CACHE or {}


def compute_lambda_rates_from_ili_for_step(step: int):
    return _lookup_lambda_rates_from_cache(step, ILI_AUTONOMOUS_WEEK_CACHE)


def _week_ili_features_for_step(lookup, step):
    if not lookup:
        return {'ili_trend': 0.0, 'ili_relative': 0.0}
    sim_week = week_index_from_step(step)
    epi_w = csv_epi_week_from_simulation_week(sim_week)
    return lookup.get(epi_w, {'ili_trend': 0.0, 'ili_relative': 0.0})


def assign_cognition_traits(nodes, rng=None) -> None:
    if rng is None:
        rng = random
    for node in nodes.values():
        grp = _node_population_group(node.get('type', ''))
        if grp == 'child':
            node['cognition_z'] = 0.0
        else:
            node['cognition_z'] = float(rng.gauss(0.0, 1.0))


def _formula_group_keys(formula_params: Dict[str, Any], group: str) -> Tuple[float, float, float]:
    if group == 'child':
        return (
            float(formula_params.get('b_child', 0.0)),
            float(formula_params.get('s_child', 1.0)),
            float(formula_params.get('phi_child', 0.5)),
        )
    if group == 'elderly':
        return (
            float(formula_params.get('b_elderly', 0.0)),
            float(formula_params.get('s_elderly', 1.0)),
            float(formula_params.get('phi_elderly', 0.5)),
        )
    return (
        float(formula_params.get('b_adult', 0.0)),
        float(formula_params.get('s_adult', 1.0)),
        float(formula_params.get('phi_adult', 0.5)),
    )


def compute_vaccination_logit_components(
    node_id: int,
    nodes: Dict[int, Any],
    month_params: Dict[str, Any],
    formula_params: Dict[str, Any],
) -> Tuple[float, Dict[str, float]]:
    node_type = nodes[node_id].get('type')
    group = _node_population_group(node_type)
    b_g, _, _ = _formula_group_keys(formula_params, group)

    policy_month_idx = int(
        month_params.get(
            'policy_month_index',
            month_params.get('month_index', 1),
        )
    )
    pe = formula_params.get('policy_effects') or []
    if not pe:
        policy_raw = 0.0
    elif policy_month_idx <= len(pe):
        policy_raw = float(pe[policy_month_idx - 1])
    else:
        policy_raw = float(pe[-1])

    beta = float(formula_params.get('beta_ili_trend', FORMULA_PARAMS.get('beta_ili_trend', 0.0)))
    eta = float(formula_params.get('eta_ili_relative', FORMULA_PARAMS.get('eta_ili_relative', 0.0)))
    gamma = float(formula_params.get('gamma_neighbor', FORMULA_PARAMS.get('gamma_neighbor', 0.0)))

    ili_trend = float(month_params.get('ili_trend', 0.0))
    ili_relative = float(month_params.get('ili_relative', 0.0))
    neighbor_benefit = _calculate_neighbor_benefit(node_id, nodes)

    ili_trend_part = beta * ili_trend
    ili_relative_part = eta * ili_relative
    peer_part = gamma * neighbor_benefit
    pol_part = policy_raw

    prior_bonus = 0.0
    if (PRIOR_VACCINATION_CONFIG or {}).get('enabled', True) and nodes[node_id].get(
        'prior_season_vaccinated'
    ):
        if group == 'child':
            prior_bonus = float(formula_params.get('prior_child', 0.0))
        elif group == 'elderly':
            prior_bonus = float(formula_params.get('prior_elderly', 0.0))
        else:
            prior_bonus = float(formula_params.get('prior_adult', 0.0))

    weighted = ili_trend_part + ili_relative_part + peer_part + pol_part
    z = b_g + weighted + prior_bonus
    components = {
        'flu': ili_trend_part + ili_relative_part,
        'peer': peer_part,
        'pol': pol_part,
        'cog': 0.0,
        'prior': prior_bonus,
        'weighted': weighted,
        'policy_raw': policy_raw,
        'ili_trend': ili_trend,
        'ili_relative': ili_relative,
        'neighbor_benefit': neighbor_benefit,
        'z': z,
    }
    return z, components


def representative_prior_vaccination_ratio(
    formula_params: Dict[str, Any],
    group: str,
) -> float:
    b_g, _, _ = _formula_group_keys(formula_params, group)
    if group == 'child':
        prior = float(formula_params.get('prior_child', 0.0))
    elif group == 'elderly':
        prior = float(formula_params.get('prior_elderly', 0.0))
    else:
        prior = float(formula_params.get('prior_adult', 0.0))
    if prior == 0.0:
        return 1.0
    p0 = _sigmoid(b_g)
    p1 = _sigmoid(b_g + prior)
    if p0 <= 1e-12:
        return float('inf') if p1 > 0 else 1.0
    return float(p1 / p0)

PRIOR_VACCINATION_RATIO_TARGETS = {
    'child': 1.877,
    'elderly': 14.6,
}


def print_prior_vaccination_ratio_diagnostic(
    formula_params: Optional[Dict[str, Any]] = None,
) -> None:
    cfg = PRIOR_VACCINATION_CONFIG or {}
    if not cfg.get('enabled', True):
        print("\nPrior-season vaccination multiplier is disabled (PRIOR_VACCINATION_CONFIG['enabled']=False)")
        return
    if formula_params is None:
        formula_params = FORMULA_PARAMS

    print("\n=== Prior-Season Vaccination: Uptake Ratio at Median Features (with / without history) ===")
    print("  Condition: H=0 (typical ILI, peer, policy, and cognition levels); P_without=sigma(b), P_with=sigma(b+kappa)")
    print(f"  {'Group':<12} {'P_without':>10} {'P_with':>10} {'Ratio':>10} {'Published OR':>12} {'prior kappa':>12}")
    print("  " + "-" * 64)

    for grp, label in (('child', 'Children'), ('elderly', 'Older adults')):
        b_g, _, _ = _formula_group_keys(formula_params, grp)
        if grp == 'child':
            kappa = float(formula_params.get('prior_child', 0.0))
        else:
            kappa = float(formula_params.get('prior_elderly', 0.0))
        p0 = _sigmoid(b_g)
        p1 = _sigmoid(b_g + kappa) if kappa != 0.0 else p0
        ratio = representative_prior_vaccination_ratio(formula_params, grp)
        target = PRIOR_VACCINATION_RATIO_TARGETS.get(grp)
        target_s = f"{target:.3f}" if target is not None else "—"
        print(
            f"  {label:<6} {p0:10.6f} {p1:10.6f} {ratio:10.3f} {target_s:>10} {kappa:12.4f}"
        )

    prior_adult = float(formula_params.get('prior_adult', 0.0))
    print(f"  Adults: prior_adult={prior_adult}; excluded from the ratio comparison.")
    print(
        "  Note: H varies by month and individual in the simulation; the table is a representative comparison."
    )


def _calculate_neighbor_benefit(node_id, nodes):
    connections = get_daytime_connections(node_id, nodes)
    if not connections:
        return 0.0

    valid_neighbors = [nid for nid in connections if nid in nodes]
    if not valid_neighbors:
        return 0.0

    vaccinated_neighbors = [nid for nid in valid_neighbors if nodes[nid].get('health_status') == 'vaccinated']
    if not vaccinated_neighbors:
        return 0.0

    unvaccinated_neighbors = [nid for nid in valid_neighbors if nodes[nid].get('health_status') != 'vaccinated']
    vac_ratio = len(vaccinated_neighbors) / len(valid_neighbors)
    if not unvaccinated_neighbors:
        return vac_ratio

    avg_v_payoff = np.mean([nodes[nid].get('individual_payoff', 0.0) for nid in vaccinated_neighbors])
    avg_u_payoff = np.mean([nodes[nid].get('individual_payoff', 0.0) for nid in unvaccinated_neighbors])
    infection_cost = COST_CONFIG.get('infection_cost', 1.0) or 1.0
    payoff_gap = max(0.0, avg_v_payoff - avg_u_payoff) / infection_cost
    return vac_ratio * payoff_gap


def policy_month_index_for_step(step: int, shift_half_months: float = 0.0) -> int:
    calendar_m = (int(step) - 1) // STEPS_PER_MONTH + 1
    if shift_half_months == 0.0:
        return calendar_m
    first_half = ((int(step) - 1) % STEPS_PER_MONTH) < HALF_MONTH_STEPS
    if shift_half_months < 0:
        if first_half:
            return calendar_m
        return min(SIMULATION_MONTHS, calendar_m + 1)
    if first_half:
        return max(1, calendar_m - 1)
    return calendar_m


def _month_params_for_step(
    step: int,
    monthly_params: Optional[Dict[int, Dict[str, Any]]],
    policy_half_month_shift: float = 0.0,
) -> Dict[str, Any]:
    month_idx = (int(step) - 1) // STEPS_PER_MONTH + 1
    mp = dict((monthly_params or {}).get(month_idx, {}))
    mp['month_index'] = month_idx
    mp['policy_month_index'] = policy_month_index_for_step(step, policy_half_month_shift)
    ili_lookup = VACCINATION_ILI_WEEKLY_LOOKUP_CACHE or {}
    wf = _week_ili_features_for_step(ili_lookup, step)
    mp['ili_trend'] = float(wf.get('ili_trend', 0.0))
    mp['ili_relative'] = float(wf.get('ili_relative', 0.0))
    return mp


def vaccination_probability_formula(node_id, nodes, month_params, formula_params):
    if nodes[node_id]['health_status'] != 'susceptible':
        return 0.0
    if nodes[node_id].get('post_vaccine_susceptible'):
        return 0.0
    z, _ = compute_vaccination_logit_components(node_id, nodes, month_params, formula_params)
    return _sigmoid(z)


def simulate_infection_and_recovery(nodes, is_daytime, rng=None, month_params=None, step=1):
    if rng is None:
        rng = random

    infection_candidates = [
        node_id for node_id, node_info in nodes.items()
        if _can_receive_infection(nodes, node_id, step)
    ]

    lam_rates = compute_lambda_rates_from_ili_for_step(step)

    for node_id in infection_candidates:
        p_nb = calculate_neighbor_infection_probability(node_id, nodes, month_params, step)
        r_auto = autonomous_infection_rate_for_node(node_id, nodes, month_params, lam_rates=lam_rates)
        r_auto = max(0.0, min(1.0, r_auto))
        infection_prob = 1.0 - (1.0 - p_nb) * (1.0 - r_auto)
        infection_prob *= _vaccine_infection_multiplier(nodes, node_id, step)
        infection_prob = max(0.0, min(1.0, infection_prob))
        if rng.random() < infection_prob:
            if nodes[node_id]['health_status'] == 'vaccinated':
                nodes[node_id].pop('vacc_antibody_step', None)
            nodes[node_id]['health_status'] = 'infected'

    rec_stats = {'total': 0, 'child': 0, 'adult': 0, 'elderly': 0}

    infected_nodes = [node_id for node_id, node_info in nodes.items()
                      if node_info['health_status'] == 'infected']

    for node_id in infected_nodes:
        node_type = nodes[node_id]['type']

        if node_type == 'child':
            recovery_rate = RECOVERY_CONFIG['children_recovery_rate']
        elif node_type in ['parent', 'single']:
            recovery_rate = RECOVERY_CONFIG['adults_recovery_rate']
        elif node_type == 'elderly':
            recovery_rate = RECOVERY_CONFIG['elderly_recovery_rate']
        else:
            recovery_rate = 0.10

        if rng.random() < recovery_rate:
            nodes[node_id]['health_status'] = 'susceptible'

            if nodes[node_id].get('ever_vaccinated'):
                nodes[node_id]['post_vaccine_susceptible'] = True
            else:
                nodes[node_id]['post_vaccine_susceptible'] = False
            rec_stats['total'] += 1
            if node_type == 'child':
                rec_stats['child'] += 1
            elif node_type in ('parent', 'single'):
                rec_stats['adult'] += 1
            elif node_type == 'elderly':
                rec_stats['elderly'] += 1

    return rec_stats


def simulate_vaccination(nodes, rng=None, month_params=None, formula_params=None, step: int = 1):
    if rng is None:
        rng = random

    if formula_params is None:
        raise ValueError("The new vaccination rule is enabled, but fitted parameters were not loaded")

    dmin = int(VACCINE_CONFIG['antibody_delay_steps_min'])
    dmax = int(VACCINE_CONFIG['antibody_delay_steps_max'])
    if dmax < dmin:
        dmin, dmax = dmax, dmin

    new_counts = {'child': 0, 'adult': 0, 'elderly': 0}
    susceptible_nodes = [
        node_id for node_id, node_info in nodes.items()
        if node_info['health_status'] == 'susceptible' and not node_info.get('post_vaccine_susceptible')
    ]
    for node_id in susceptible_nodes:
        p_vac = vaccination_probability_formula(node_id, nodes, month_params or {}, formula_params)
        if rng.random() < p_vac:
            nodes[node_id]['health_status'] = 'vaccinated'
            nodes[node_id]['ever_vaccinated'] = True
            nodes[node_id]['post_vaccine_susceptible'] = False
            delay = rng.randint(dmin, dmax)
            nodes[node_id]['vacc_antibody_step'] = int(step) + int(delay)
            t = nodes[node_id]['type']
            if t == 'child':
                new_counts['child'] += 1
            elif t in ('parent', 'single'):
                new_counts['adult'] += 1
            elif t == 'elderly':
                new_counts['elderly'] += 1

    return new_counts


def run_epidemic_simulation(nodes, num_steps=100, display_interval=5, verbose=False, rng=None, sample_steps=None,
                            monthly_params=None, policy_half_month_shift=None,
                            formula_params=None):
    if rng is None:
        rng = random

    global VACCINATION_FORMULA_PARAMS_CACHE, VACCINATION_ILI_WEEKLY_LOOKUP_CACHE
    if formula_params is not None:
        fp = formula_params
    else:
        if VACCINATION_FORMULA_PARAMS_CACHE is None:
            VACCINATION_FORMULA_PARAMS_CACHE = FORMULA_PARAMS
        fp = VACCINATION_FORMULA_PARAMS_CACHE
    if fp is None:
        raise ValueError("Failed to enable the new vaccination rule: vaccination parameters were not loaded")
    if VACCINATION_ILI_WEEKLY_LOOKUP_CACHE is None:
        load_weekly_ili_caches()

    if policy_half_month_shift is None:
        policy_half_month_shift = float(SIMULATION_CONFIG.get('policy_half_month_shift', 0.0))
    else:
        policy_half_month_shift = float(policy_half_month_shift)

    if sample_steps is not None:
        if isinstance(sample_steps, (list, tuple, set)):
            sample_set = set(sample_steps)
            sample_order = sorted(sample_set)
        else:
            raise ValueError("sample_steps must be iterable")
    else:
        sample_set = None
        sample_order = None

    time_series_data = []
    cum_rec_total = cum_rec_child = cum_rec_adult = cum_rec_elderly = 0

    for step in range(1, num_steps + 1):
        month_idx = (step - 1) // STEPS_PER_MONTH + 1
        month_params = monthly_params.get(month_idx, {}) if monthly_params else {}
        month_params = dict(month_params)
        month_params['month_index'] = month_idx
        month_params['policy_month_index'] = policy_month_index_for_step(
            step, policy_half_month_shift
        )
        ili_lookup = VACCINATION_ILI_WEEKLY_LOOKUP_CACHE or {}
        wf = _week_ili_features_for_step(ili_lookup, step)
        month_params['ili_trend'] = float(wf.get('ili_trend', 0.0))
        month_params['ili_relative'] = float(wf.get('ili_relative', 0.0))

        is_daytime = (step % 2 == 1)
        time_period = "Daytime" if is_daytime else "Nighttime"

        if is_daytime:
            calculate_payoffs(nodes, is_daytime, month_params, step)

        if not is_daytime:
            vac_new = simulate_vaccination(
                nodes, rng=rng, month_params=month_params, formula_params=fp, step=step
            )
        else:
            vac_new = {'child': 0, 'adult': 0, 'elderly': 0}

        rec_st = simulate_infection_and_recovery(
            nodes, is_daytime, rng=rng, month_params=month_params, step=step
        )
        cum_rec_total += rec_st['total']
        cum_rec_child += rec_st['child']
        cum_rec_adult += rec_st['adult']
        cum_rec_elderly += rec_st['elderly']

        susceptible_count = sum(1 for node in nodes.values() if node['health_status'] == 'susceptible')
        infected_count = sum(1 for node in nodes.values() if node['health_status'] == 'infected')
        vaccinated_count = sum(1 for node in nodes.values() if node['health_status'] == 'vaccinated')
        recovered_count = cum_rec_total

        child_vaccinated = sum(1 for node in nodes.values()
                               if node['type'] == 'child' and node['health_status'] == 'vaccinated')
        child_infected = sum(1 for node in nodes.values()
                             if node['type'] == 'child' and node['health_status'] == 'infected')
        child_recovered = cum_rec_child

        adult_vaccinated = sum(1 for node in nodes.values()
                               if node['type'] in ['parent', 'single'] and node['health_status'] == 'vaccinated')
        adult_infected = sum(1 for node in nodes.values()
                             if node['type'] in ['parent', 'single'] and node['health_status'] == 'infected')
        adult_recovered = cum_rec_adult

        elderly_vaccinated = sum(1 for node in nodes.values()
                                 if node['type'] == 'elderly' and node['health_status'] == 'vaccinated')
        elderly_infected = sum(1 for node in nodes.values()
                               if node['type'] == 'elderly' and node['health_status'] == 'infected')
        elderly_recovered = cum_rec_elderly

        school_infected = child_infected if is_daytime else 0

        company_infected = adult_infected if is_daytime else 0

        outdoor_infected = elderly_infected if is_daytime else 0

        if not is_daytime:
            family_infected = sum(1 for node in nodes.values()
                                  if node.get('family_id') is not None
                                  and node['health_status'] == 'infected')
        else:
            family_infected = 0

        if sample_set is None or step in sample_set:
            time_series_data.append({
                'step': step,
                'susceptible': susceptible_count,
                'infected': infected_count,
                'vaccinated': vaccinated_count,
                'recovered': recovered_count,

                'child_vaccinated': child_vaccinated,
                'adult_vaccinated': adult_vaccinated,
                'elderly_vaccinated': elderly_vaccinated,

                'child_vaccinated_new': vac_new['child'],
                'adult_vaccinated_new': vac_new['adult'],
                'elderly_vaccinated_new': vac_new['elderly'],
                'child_infected_recovered': child_infected + child_recovered,
                'adult_infected_recovered': adult_infected + adult_recovered,
                'elderly_infected_recovered': elderly_infected + elderly_recovered,

                'school_infected': school_infected,
                'company_infected': company_infected,
                'outdoor_infected': outdoor_infected,
                'family_infected': family_infected
            })

    if sample_order is None:
        return time_series_data

    ordered_data = []
    data_by_step = {entry['step']: entry for entry in time_series_data}
    for step in sample_order:
        if step in data_by_step:
            ordered_data.append(data_by_step[step])
    return ordered_data


def run_single_simulation_wrapper(args):
    formula_params_override = None
    if len(args) >= 7:
        (
            run_index,
            nodes_template,
            num_steps,
            monthly_params,
            seed_base,
            policy_half_month_shift,
            formula_params_override,
        ) = args[:7]
    elif len(args) >= 6:
        run_index, nodes_template, num_steps, monthly_params, seed_base, policy_half_month_shift = args[:6]
    else:
        run_index, nodes_template, num_steps, monthly_params, seed_base = args
        policy_half_month_shift = float(SIMULATION_CONFIG.get('policy_half_month_shift', 0.0))

    simulation_rng = random.Random(seed_base)

    nodes_copy = {}
    for node_id, node_info in nodes_template.items():
        nodes_copy[node_id] = node_info.copy()
        nodes_copy[node_id]['health_status'] = None

    nodes_copy, _ = assign_health_status(nodes_copy, rng=simulation_rng)

    time_series_data = run_epidemic_simulation(
        nodes_copy,
        num_steps=num_steps,
        display_interval=5,
        verbose=False,
        rng=simulation_rng,
        monthly_params=monthly_params,
        policy_half_month_shift=policy_half_month_shift,
        formula_params=formula_params_override,
    )

    return (run_index, time_series_data)


def _daily_new_vaccination_dicts_one_run(ts_data: List[Dict[str, Any]]) -> Tuple[Dict[int, float], Dict[int, float], Dict[int, float]]:
    dc: Dict[int, float] = {}
    da: Dict[int, float] = {}
    de: Dict[int, float] = {}
    for entry in ts_data:
        s = int(entry['step'])
        day = (s + 1) // 2
        dc[day] = dc.get(day, 0.0) + float(entry.get('child_vaccinated_new', 0))
        da[day] = da.get(day, 0.0) + float(entry.get('adult_vaccinated_new', 0))
        de[day] = de.get(day, 0.0) + float(entry.get('elderly_vaccinated_new', 0))
    return dc, da, de


def _monthly_vaccination_from_daily_dicts(
    dc: Dict[int, float],
    da: Dict[int, float],
    de: Dict[int, float],
) -> Dict[int, Tuple[float, float, float]]:
    out: Dict[int, Tuple[float, float, float]] = {}
    for m in range(1, SIMULATION_MONTHS + 1):
        d0 = (m - 1) * DAYS_PER_MONTH + 1
        d1 = m * DAYS_PER_MONTH
        out[m] = (
            sum(dc.get(d, 0.0) for d in range(d0, d1 + 1)),
            sum(da.get(d, 0.0) for d in range(d0, d1 + 1)),
            sum(de.get(d, 0.0) for d in range(d0, d1 + 1)),
        )
    return out


def _t_critical_95(n: int) -> float:
    if n <= 1:
        return 0.0
    if n >= 30:
        return 1.96
    if HAS_SCIPY:
        return float(stats.t.ppf(0.975, n - 1))
    if n == 20:
        return 2.093
    return 1.96


def _mean_and_ci_95(values: Sequence[float]) -> Tuple[float, float, float]:
    arr = np.asarray(values, dtype=float)
    n = len(arr)
    if n == 0:
        return 0.0, 0.0, 0.0
    mean = float(np.mean(arr))
    if n < 2:
        return mean, mean, mean
    std = float(np.std(arr, ddof=1))
    margin = _t_critical_95(n) * std / np.sqrt(n)
    return mean, mean - margin, mean + margin


def _vac_increment_ci_fields(
    child_vals: Sequence[float],
    adult_vals: Sequence[float],
    elderly_vals: Sequence[float],
) -> Dict[str, float]:
    gc, gc_lo, gc_hi = _mean_and_ci_95(child_vals)
    ga, ga_lo, ga_hi = _mean_and_ci_95(adult_vals)
    ge, ge_lo, ge_hi = _mean_and_ci_95(elderly_vals)
    return {
        'child_vaccinated_new': max(0.0, gc),
        'child_vaccinated_new_ci_lower': max(0.0, gc_lo),
        'child_vaccinated_new_ci_upper': max(0.0, gc_hi),
        'adult_vaccinated_new': max(0.0, ga),
        'adult_vaccinated_new_ci_lower': max(0.0, ga_lo),
        'adult_vaccinated_new_ci_upper': max(0.0, ga_hi),
        'elderly_vaccinated_new': max(0.0, ge),
        'elderly_vaccinated_new_ci_lower': max(0.0, ge_lo),
        'elderly_vaccinated_new_ci_upper': max(0.0, ge_hi),
    }


def calculate_monthly_increments(all_time_series_data, num_runs=1, initial_summary=None):
    month_end_steps = list(range(STEPS_PER_MONTH, DEFAULT_SIMULATION_STEPS + 1, STEPS_PER_MONTH))
    if len(SIMULATION_MONTH_LABELS) == SIMULATION_MONTHS:
        month_names = list(SIMULATION_MONTH_LABELS)
    else:
        month_names = [f'Month {i}' for i in range(1, SIMULATION_MONTHS + 1)]

    has_vac_flow = any(
        'child_vaccinated_new' in e for ts in all_time_series_data for e in ts
    )

    vac_mean_month: Dict[int, Tuple[float, float, float]] = {}
    vac_per_run_monthly: List[Dict[int, Tuple[float, float, float]]] = []
    if has_vac_flow:
        for ts in all_time_series_data:
            dc, da, de = _daily_new_vaccination_dicts_one_run(ts)
            vac_per_run_monthly.append(_monthly_vaccination_from_daily_dicts(dc, da, de))
        for m in range(1, SIMULATION_MONTHS + 1):
            cs = [pr[m][0] for pr in vac_per_run_monthly]
            ads = [pr[m][1] for pr in vac_per_run_monthly]
            els = [pr[m][2] for pr in vac_per_run_monthly]
            vac_mean_month[m] = (float(np.mean(cs)), float(np.mean(ads)), float(np.mean(els)))

    initial_total_infected = 0
    if initial_summary is not None:
        initial_total_infected = initial_summary.get('infected', 0) + initial_summary.get('recovered', 0)

    all_steps = set()
    for data in all_time_series_data:
        all_steps.update([d['step'] for d in data])

    monthly_data = {}
    for month_idx, step in enumerate(month_end_steps, 1):
        if step not in all_steps:
            continue

        child_vac_values = []
        adult_vac_values = []
        elderly_vac_values = []
        total_infected_values = []

        for time_series_data in all_time_series_data:
            data_by_step = {d['step']: d for d in time_series_data}
            if step in data_by_step:
                data = data_by_step[step]
                child_vac_values.append(data.get('child_vaccinated', 0))
                adult_vac_values.append(data.get('adult_vaccinated', 0))
                elderly_vac_values.append(data.get('elderly_vaccinated', 0))
                infected_count = data.get('infected', 0)
                total_infected_values.append(infected_count)

        if child_vac_values:
            monthly_data[month_idx] = {
                'month_name': month_names[month_idx - 1],
                'child_vaccinated': np.mean(child_vac_values),
                'adult_vaccinated': np.mean(adult_vac_values),
                'elderly_vaccinated': np.mean(elderly_vac_values),
                'total_infected': np.mean(total_infected_values),
            }

    monthly_increments: Dict[int, Dict[str, Any]] = {}

    if has_vac_flow:
        infected_prev = initial_total_infected
        for m in range(1, SIMULATION_MONTHS + 1):
            cs = [pr[m][0] for pr in vac_per_run_monthly]
            ads = [pr[m][1] for pr in vac_per_run_monthly]
            els = [pr[m][2] for pr in vac_per_run_monthly]
            name = month_names[m - 1]
            if m in monthly_data:
                inf_row = monthly_data[m]['total_infected']
                tin = max(0.0, float(inf_row - infected_prev))
                infected_prev = float(inf_row)
            else:
                tin = 0.0
            row: Dict[str, Any] = {
                'month_name': name,
                'total_infected_new': tin,
            }
            row.update(_vac_increment_ci_fields(cs, ads, els))
            monthly_increments[m] = row
        return monthly_increments

    prev_child_vac = 0
    prev_adult_vac = 0
    prev_elderly_vac = 0
    prev_total_infected = initial_total_infected

    for month_idx in sorted(monthly_data.keys()):
        data = monthly_data[month_idx]
        child_vac_new = data['child_vaccinated'] - prev_child_vac
        adult_vac_new = data['adult_vaccinated'] - prev_adult_vac
        elderly_vac_new = data['elderly_vaccinated'] - prev_elderly_vac
        total_infected_new = data['total_infected'] - prev_total_infected

        monthly_increments[month_idx] = {
            'month_name': data['month_name'],
            'child_vaccinated_new': max(0, child_vac_new),
            'adult_vaccinated_new': max(0, adult_vac_new),
            'elderly_vaccinated_new': max(0, elderly_vac_new),
            'total_infected_new': max(0, total_infected_new),
        }

        prev_child_vac = data['child_vaccinated']
        prev_adult_vac = data['adult_vaccinated']
        prev_elderly_vac = data['elderly_vaccinated']
        prev_total_infected = data['total_infected']

    return monthly_increments


def _fmt_mean_ci(data: Dict[str, Any], prefix: str) -> str:
    mean = data[prefix]
    lo_key, hi_key = f"{prefix}_ci_lower", f"{prefix}_ci_upper"
    if lo_key in data and hi_key in data:
        return f"{mean:.2f} [{data[lo_key]:.2f}, {data[hi_key]:.2f}]"
    return f"{mean:.2f}"


def print_monthly_increments(monthly_increments):
    has_vac_ci = any('child_vaccinated_new_ci_lower' in d for d in monthly_increments.values())

    print("\n" + "=" * 100)
    print("Monthly New Vaccinations and Infections")
    print(
        "Note: vaccination columns sum calendar-day increments, calculated with the same rule as the daily "
        f"new-vaccination curve, over each {DAYS_PER_MONTH}-day month and then average across runs. "
        "This equals summing the plotted daily means by month. Legacy data without per-step "
        "*_vaccinated_new fields still use differences in month-end vaccination prevalence."
    )
    if has_vac_ci:
        print("Vaccination-column format: mean [95% CI lower, upper] using a t distribution across runs, consistent with the plots.")
    print("=" * 100)

    col_w = 28 if has_vac_ci else 15
    print(
        f"\n{'Month':<10} {'New child vaccinations':<{col_w}} {'New adult vaccinations':<{col_w}} "
        f"{'New older-adult vaccinations':<{col_w}} {'Total new infections':<20}"
    )
    print("-" * (10 + col_w * 3 + 15))

    for month_idx in sorted(monthly_increments.keys()):
        data = monthly_increments[month_idx]
        print(
            f"{data['month_name']:<10} "
            f"{_fmt_mean_ci(data, 'child_vaccinated_new'):<{col_w}} "
            f"{_fmt_mean_ci(data, 'adult_vaccinated_new'):<{col_w}} "
            f"{_fmt_mean_ci(data, 'elderly_vaccinated_new'):<{col_w}} "
            f"{data['total_infected_new']:<15.2f}"
        )

    print("\nDetails:")
    for month_idx in sorted(monthly_increments.keys()):
        data = monthly_increments[month_idx]
        print(f"\n{data['month_name']}:")
        print(f"  New child vaccinations: {_fmt_mean_ci(data, 'child_vaccinated_new')} people")
        print(f"  New adult vaccinations: {_fmt_mean_ci(data, 'adult_vaccinated_new')} people")
        print(f"  New older-adult vaccinations: {_fmt_mean_ci(data, 'elderly_vaccinated_new')} people")
        print(f"  Total new infections: {data['total_infected_new']:.2f} people")

    print("=" * 100)

MONTHLY_VAC_REAL_VALUES = {
    'child': [3, 15.6, 31.4, 27.9, 21.8, 2.4, 0.6, 0.2],
    'adult': [0.5, 3.8, 7.4, 6.0, 4.3, 0.3, 0.07, 0.03],
    'elderly': [0.05, 0.97, 2.16, 1.67, 0.87, 0.05, 0.02, 0.02],
}


def _format_monthly_annotation(value: float) -> str:
    v = float(value)
    if abs(v - round(v)) < 1e-9:
        iv = int(round(v))
        if abs(iv) >= 1:
            return str(iv)
    av = abs(v)
    if av >= 1:
        s = f"{v:.1f}".rstrip('0').rstrip('.')
        return s
    if av >= 0.01:
        s = f"{v:.2f}".rstrip('0').rstrip('.')
        return s
    return f"{v:.2g}"


def _save_figure_to_output(filename: str) -> str:
    output_dir = OUTPUT_CONFIG['output_dir']
    if output_dir:
        output_dir = os.path.normpath(output_dir)
        try:
            os.makedirs(output_dir, exist_ok=True)
            image_path = os.path.join(output_dir, filename)
        except OSError as e:
            print(f"Error: could not create output directory '{output_dir}': {e}")
            image_path = filename
    else:
        image_path = filename
    plt.savefig(image_path, dpi=300, bbox_inches='tight')
    return os.path.abspath(image_path)


def plot_monthly_vaccination_ci_by_group(
    monthly_increments: Dict[int, Dict[str, Any]],
    num_runs: int = 1,
    real_values: Optional[Dict[str, Sequence[float]]] = None,
) -> None:
    if real_values is None:
        real_values = MONTHLY_VAC_REAL_VALUES

    month_indices = sorted(monthly_increments.keys())
    if not month_indices:
        print("Warning: no monthly vaccination data; skipping group-specific monthly confidence-interval plots.")
        return

    x_labels = [monthly_increments[m]['month_name'] for m in month_indices]
    x_pos = np.arange(len(month_indices))
    bar_width = 0.55

    group_specs = [
        (
            'child',
            'Children',
            'child_vaccinated_new',
            'orange',
            'children_monthly_new_vaccination_ci.png',
        ),
        (
            'adult',
            'Adults',
            'adult_vaccinated_new',
            'blue',
            'adults_monthly_new_vaccination_ci.png',
        ),
        (
            'elderly',
            'Older adults',
            'elderly_vaccinated_new',
            'red',
            'older_adults_monthly_new_vaccination_ci.png',
        ),
    ]

    has_vac_ci = any(
        'child_vaccinated_new_ci_lower' in monthly_increments[m] for m in month_indices
    )
    ci_note = (
        f"95% CI from {num_runs} simulations vs observed values"
        if has_vac_ci and num_runs > 1
        else "Monthly new vaccinations vs observed values"
    )

    print("\n" + "=" * 80)
    print("Whether monthly observed values fall within the simulated 95% confidence intervals")
    print("=" * 80)

    for group_key, group_name, prefix, color, filename in group_specs:
        lo_key, hi_key = f"{prefix}_ci_lower", f"{prefix}_ci_upper"
        if has_vac_ci:
            lowers = [float(monthly_increments[m][lo_key]) for m in month_indices]
            uppers = [float(monthly_increments[m][hi_key]) for m in month_indices]
        else:
            means = [float(monthly_increments[m][prefix]) for m in month_indices]
            lowers = means
            uppers = means

        heights = [max(0.0, u - l) for l, u in zip(lowers, uppers)]

        rv_list = list(real_values.get(group_key, []))
        if len(rv_list) < len(month_indices):
            print(
                f"Warning: only {len(rv_list)} observed values are available for {group_name}; "
                f"{len(month_indices)} are required. Missing months will not show observed points."
            )
        elif len(rv_list) > len(month_indices):
            rv_list = rv_list[: len(month_indices)]

        plt.figure(figsize=(10, 6))
        plt.bar(
            x_pos,
            heights,
            width=bar_width,
            bottom=lowers,
            color=color,
            alpha=0.35,
            edgecolor=color,
            linewidth=1.2,
            align='center',
            label='Simulated 95% CI',
            zorder=1,
        )

        in_x, in_y = [], []
        out_x, out_y = [], []
        print(f"\n[{group_name}]")
        for i, m_idx in enumerate(month_indices):
            if i >= len(rv_list):
                continue
            rv = float(rv_list[i])
            lo, hi = lowers[i], uppers[i]
            inside = lo <= rv <= hi if has_vac_ci else True
            status = "inside CI" if inside else "outside CI"
            print(
                f"  {x_labels[i]}: observed={rv:.4g}, CI=[{lo:.4g}, {hi:.4g}] -> {status}"
            )
            if inside:
                in_x.append(x_pos[i])
                in_y.append(rv)
            else:
                out_x.append(x_pos[i])
                out_y.append(rv)

        if in_x:
            plt.scatter(
                in_x,
                in_y,
                c='#2ca02c',
                marker='D',
                s=72,
                zorder=3,
                edgecolors='black',
                linewidths=0.8,
                label='Observed value (inside CI)',
            )
        if out_x:
            plt.scatter(
                out_x,
                out_y,
                c='#d62728',
                marker='D',
                s=72,
                zorder=3,
                edgecolors='black',
                linewidths=0.8,
                label='Observed value (outside CI)',
            )

        for i, rv in enumerate(rv_list):
            if i >= len(month_indices):
                break
            plt.annotate(
                _format_monthly_annotation(rv),
                xy=(x_pos[i], rv),
                xytext=(0, 10),
                textcoords='offset points',
                ha='center',
                va='bottom',
                fontsize=10,
                color='black',
                zorder=4,
            )

        y_all = list(lowers) + list(uppers) + rv_list
        if y_all:
            y_min, y_max = min(y_all), max(y_all)
            pad = max((y_max - y_min) * 0.12, 0.5)
            plt.ylim(bottom=max(0.0, y_min - pad), top=y_max + pad)

        plt.xticks(x_pos, x_labels)
        plt.xlabel('Month (August-March)', fontsize=12)
        plt.ylabel('Monthly new vaccinations', fontsize=12)
        plt.title(f'{group_name}: {ci_note}', fontsize=14)
        plt.legend(fontsize=9, loc='upper right')
        plt.grid(True, alpha=0.3, axis='y')
        plt.tight_layout()

        abs_path = _save_figure_to_output(filename)
        print(f"Monthly new-vaccination confidence-interval plot for {group_name} saved to: {abs_path}")
        plt.show()
        plt.close()

    print("=" * 80)


def export_monthly_vaccination_ci_csv(
    monthly_increments: Dict[int, Dict[str, Any]],
    scenario_label: Optional[str] = None,
    output_dir: Optional[str] = None,
    csv_stem: Optional[str] = None,
) -> Optional[str]:
    if output_dir is None:
        output_dir = OUTPUT_CONFIG.get('monthly_ci_csv_dir') or OUTPUT_CONFIG.get('output_dir', '')
    if not monthly_increments:
        print("Warning: no monthly vaccination data; skipping confidence-interval CSV export.")
        return None

    output_dir = os.path.normpath(output_dir)
    try:
        os.makedirs(output_dir, exist_ok=True)
    except OSError as e:
        print(f"Warning: could not create CSV output directory '{output_dir}': {e}")
        return None

    group_specs = [
        ('child', 'Children'),
        ('adult', 'Adults'),
        ('elderly', 'Older_adults'),
    ]

    rows: List[Dict[str, Any]] = []
    for m_idx in sorted(monthly_increments.keys()):
        data = monthly_increments[m_idx]
        row: Dict[str, Any] = {
            'month_idx': m_idx,
            'month_name': data.get('month_name', ''),
        }
        for prefix, label in group_specs:
            key = f'{prefix}_vaccinated_new'
            row[f'{label}_mean'] = float(data.get(key, np.nan))
            row[f'{label}_CI_lower'] = float(data.get(f'{key}_ci_lower', np.nan))
            row[f'{label}_CI_upper'] = float(data.get(f'{key}_ci_upper', np.nan))
        rows.append(row)

    df = pd.DataFrame(rows)
    if csv_stem:
        stem = csv_stem
    else:
        stem = 'monthly_vaccination_confidence_intervals'
        if scenario_label:
            stem = f'{stem}_{scenario_label}'
    csv_path = os.path.join(output_dir, f'{stem}.csv')
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    abs_path = os.path.abspath(csv_path)
    print(f"Monthly vaccination confidence-interval data saved to: {abs_path}")
    return abs_path


def export_to_excel(time_series_data, filename=None, sheet_name=None):
    if filename is None:
        filename = OUTPUT_CONFIG['data_filename']

    output_dir = OUTPUT_CONFIG['output_dir']

    if output_dir:
        output_dir = os.path.normpath(output_dir)
        try:
            os.makedirs(output_dir, exist_ok=True)
            file_path = os.path.join(output_dir, filename)
        except OSError as e:
            print(f"Error: could not create output directory '{output_dir}': {e}")
            print("Saving to the current directory")
            file_path = filename
    else:
        file_path = filename

    data = {
        'Time step': [data['step'] for data in time_series_data],
        'Infected population': [data['infected'] for data in time_series_data],
        'Vaccinated population': [data['vaccinated'] for data in time_series_data],
        'Recovered population': [data['recovered'] for data in time_series_data]
    }

    df = pd.DataFrame(data)

    try:
        file_exists = os.path.exists(file_path)

        if file_exists:
            from openpyxl import load_workbook

            if sheet_name is None:
                wb = load_workbook(file_path)
                existing_sheets = wb.sheetnames
                wb.close()

                max_run = 0
                for sheet in existing_sheets:
                    if sheet.startswith('Run'):
                        try:
                            run_num = int(sheet.replace('Run', ''))
                            max_run = max(max_run, run_num)
                        except ValueError:
                            pass

                sheet_name = f'Run{max_run + 1}'
            else:
                wb = load_workbook(file_path)
                existing_sheets = wb.sheetnames
                wb.close()

                original_name = sheet_name
                counter = 1
                while sheet_name in existing_sheets:
                    sheet_name = f'{original_name}_{counter}'
                    counter += 1

            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            abs_file_path = os.path.abspath(file_path)
            print(f"Simulation results appended to worksheet '{sheet_name}': {abs_file_path}")
        else:
            if sheet_name is None:
                sheet_name = 'Run1'

            df.to_excel(file_path, sheet_name=sheet_name, index=False, engine='openpyxl')
            abs_file_path = os.path.abspath(file_path)
            print(f"Simulation results exported to worksheet '{sheet_name}' in a new file: {abs_file_path}")
    except (ModuleNotFoundError, ImportError):
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        csv_filename = file_path.replace('.xlsx', f'_{timestamp}.csv')
        df.to_csv(csv_filename, index=False, encoding='utf-8-sig')
        abs_file_path = os.path.abspath(csv_filename)
        print(f"Simulation results exported to: {abs_file_path} (CSV format, compatible with Excel)")
        print("Tip: to export directly to a multi-worksheet Excel file, run: pip install openpyxl")


def calculate_family_distribution():
    A_count = 121
    B_count = 121
    C_count = 157
    D_count = 170
    E_count = 231
    F_count = 99
    G_count = 98
    H_count = 42

    children_used = (E_count * 1 + F_count * 2 + G_count * 1 + H_count * 2)

    parents_used = (C_count * 2 + E_count * 2 + F_count * 2 + G_count * 2 + H_count * 2)

    singles_used = A_count * 1

    elderly_used = (B_count * 1 + D_count * 2 + G_count * 1 + H_count * 1)

    total_families = A_count + B_count + C_count + D_count + E_count + F_count + G_count + H_count
    total_population = (A_count * 1 + B_count * 1 + C_count * 2 + D_count * 2 +
                        E_count * 3 + F_count * 4 + G_count * 4 + H_count * 5)
    avg_family_size = total_population / total_families if total_families > 0 else 0

    expected_children = POPULATION_CONFIG['children_count']
    expected_adults = POPULATION_CONFIG['adults_count']
    expected_elderly = POPULATION_CONFIG['elderly_count']
    expected_total = expected_children + expected_adults + expected_elderly

    parent_pairs = parents_used // 2
    singles = singles_used

    result = {
        'A': A_count,
        'B': B_count,
        'C': C_count,
        'D': D_count,
        'E': E_count,
        'F': F_count,
        'G': G_count,
        'H': H_count,
        'total_families': total_families,
        'total_population': total_population,
        'expected_population': expected_total,
        'avg_family_size': avg_family_size,
        'children_used': children_used,
        'expected_children': expected_children,
        'parents_used': parents_used,
        'parent_pairs': parent_pairs,
        'singles_used': singles_used,
        'elderly_used': elderly_used,
        'expected_elderly': expected_elderly,
        'expected_adults': expected_adults
    }

    return result


def print_family_distribution(family_dist):
    print("\n" + "=" * 60)
    print("Household-Type Distribution")
    print("=" * 60)

    print(f"\nSingle-person households:")
    print(f"  Type A (single adult, 1 person): {family_dist['A']} households")
    print(f"  Type B (older adult living alone, 1 person): {family_dist['B']} households")

    print(f"\nNuclear families without children:")
    print(f"  Type C (adult couple, 2 people): {family_dist['C']} households")
    print(f"  Type D (older couple, 2 people): {family_dist['D']} households")

    print(f"\nNuclear families with children:")
    print(f"  Type E (parents + 1 child, 3 people): {family_dist['E']} households")
    print(f"  Type F (parents + 2 children, 4 people): {family_dist['F']} households")
    print(f"  Type G (parents + 1 child + 1 older adult, 4 people): {family_dist['G']} households")
    print(f"  Type H (parents + 2 children + 1 older adult, 5 people): {family_dist['H']} households")

    print(f"\nTotals:")
    print(f"  Total households: {family_dist['total_families']}")
    print(f"  Total population: {family_dist['total_population']} people")
    print(f"  Expected population: {family_dist['expected_population']} people")
    if abs(family_dist['total_population'] - family_dist['expected_population']) > 1:
        print(f"  Warning: total population mismatch; difference: {family_dist['total_population'] - family_dist['expected_population']} people")
    print(f"  Average household size: {family_dist['avg_family_size']:.2f} people per household")

    print(f"\nResource use:")
    print(f"  Children used: {family_dist['children_used']} people (expected: {family_dist['expected_children']})")
    if abs(family_dist['children_used'] - family_dist['expected_children']) > 0:
        print(f"  Warning: child-count difference: {family_dist['children_used'] - family_dist['expected_children']} people")

    print(f"  Parents used: {family_dist['parents_used']} people ({family_dist['parent_pairs']} pairs)")
    print(f"  Single adults used: {family_dist['singles_used']} people")
    print(
        f"  Total adults used: {family_dist['parents_used'] + family_dist['singles_used']} people (expected: {family_dist['expected_adults']})")
    if abs((family_dist['parents_used'] + family_dist['singles_used']) - family_dist['expected_adults']) > 0:
        print(
            f"  Warning: adult-count difference: {(family_dist['parents_used'] + family_dist['singles_used']) - family_dist['expected_adults']} people")

    print(f"  Older adults used: {family_dist['elderly_used']} people (expected: {family_dist['expected_elderly']})")
    if abs(family_dist['elderly_used'] - family_dist['expected_elderly']) > 0:
        print(f"  Warning: older-adult-count difference: {family_dist['elderly_used'] - family_dist['expected_elderly']} people")

POPULATION_CONFIG = {
    'children_count': 611,
    'adults_count': 1375,
    'elderly_count': 601,
}

NETWORK_CONFIG = {

    'school_categories': [
        {
            'name': 'Small school',
            'count': 3,
            'student_ratio': 0.117,
            'connections': (2, 4),
        },
        {
            'name': 'Medium school',
            'count': 6,
            'student_ratio': 0.475,
            'connections': (3, 5),
        },
        {
            'name': 'Large school',
            'count': 3,
            'student_ratio': 0.408,
            'connections': (4, 7),
        },
    ],
    'num_work_groups': 100,
    'adults_connections': (2, 3),
    'elderly_connections': (2, 3),
}

INFECTION_CONFIG = {
    'p_single_infection': 0.05,
}

RECOVERY_CONFIG = {
    'children_recovery_rate': 0.07,
    'adults_recovery_rate': 0.10,
    'elderly_recovery_rate': 0.035,
}

VACCINE_CONFIG = {
    'antibody_delay_steps_min': 30,
    'antibody_delay_steps_max': 60,
    've_child': 0.486,
    've_adult': 0.367,
    've_elderly': 0.49,
}

PRIOR_VACCINATION_CONFIG = {
    'enabled': True,
    'coverage_child': 0.168,
    'coverage_adult': 0.0163,
    'coverage_elderly': 0.0097,
}

COST_CONFIG = {
    'infection_cost': 10000,
    'vaccination_cost': 10,
    'recovery_cost': 0,
}

ILI_LAMBDA_DEFAULT_ATTACK_CHILD = 0.127
ILI_LAMBDA_DEFAULT_ATTACK_ADULT = 0.044
ILI_LAMBDA_DEFAULT_ATTACK_ELDERLY = 0.072

ILI_LAMBDA_FORMULA_CONFIG = {
    "pi_child": 0.15,
    "pi_adult": 0.70,
    "pi_elderly": 0.15,
    "attack_child": ILI_LAMBDA_DEFAULT_ATTACK_CHILD,
    "attack_adult": ILI_LAMBDA_DEFAULT_ATTACK_ADULT,
    "attack_elderly": ILI_LAMBDA_DEFAULT_ATTACK_ELDERLY,
    "h_child": 0.45,
    "h_adult": 0.25,
    "h_elderly": 0.40,
    "calibration_c": 1,
}

ILI_CONFIG = {
    'weekly_csv': os.path.join(PROJECT_DIR, 'data', 'weekly_ili_2026_forecast_inputs.csv'),
    'default_weekly_positivity': 0.25,
}

VACCINATION_FORMULA_PARAMS_CACHE = None
VACCINATION_ILI_WEEKLY_LOOKUP_CACHE = None
ILI_AUTONOMOUS_WEEK_CACHE = None
ILI_WEEKLY_CSV_LOAD_FAILED = False

SIMULATION_CONFIG = {
    'num_steps': DEFAULT_SIMULATION_STEPS,
    'network_seed': 42,
}

MONTH_BOUNDARIES = [
    (1, 60),
    (61, 120),
    (121, 180),
    (181, 240),
    (241, 300),
    (301, 360),
    (361, 420),
    (421, 480),
]

MONTHLY_PARAMS = {
    1: {
        'p_single_infection': 0.01,
    },
    2: {
        'p_single_infection': 0.01,
    },
    3: {
        'p_single_infection': 0.01,
    },
    4: {
        'p_single_infection': 0.01,
    },
    5: {
        'p_single_infection': 0.01,
    },
    6: {
        'p_single_infection': 0.01,
    },
    7: {
        'p_single_infection': 0.01,
    },
    8: {
        'p_single_infection': 0.01,
    },
}

FORMULA_PARAMS = {
    "b_child": -2.352,
    "b_adult": -4.698,
    "b_elderly": -4.9,
    "beta_ili_trend": 0.721,
    "eta_ili_relative": -1.496,
    "gamma_neighbor": -1.041,
    "prior_child": 0.64,
    "prior_adult": 0.0,
    "prior_elderly": 4.881,
    "policy_effects": [
        -6.0,
        -4.293,
        -3.427,
        -2.993,
        -2.718,
        -4.872,
        -5.998,
        -6.216,
    ],
}

OUTPUT_CONFIG = {
    'output_dir': os.path.join(PROJECT_DIR, 'outputs', 'simulation'),
    'monthly_ci_csv_dir': os.path.join(PROJECT_DIR, 'outputs', 'monthly_ci'),

    'article_csv_dir': os.path.join(PROJECT_DIR, 'outputs', 'article_csv'),
    'data_filename': 'influenza_transmission_simulation_results.xlsx',

    'export_monthly_plots': True,
    'export_excel': False,
}

FORECAST_2026_SCENARIO_SPECS = [
    {'id': 'baseline', 'label': 'baseline', 'policy_shift': 0.0, 'policy_factor': 1.0, 'csv_stem': '2026_baseline'},
    {'id': 'advance', 'label': 'policy_advance_2w', 'policy_shift': -0.5, 'policy_factor': 1.0, 'csv_stem': '2026_advance_2w'},
    {'id': 'delay', 'label': 'policy_delay_2w', 'policy_shift': 0.5, 'policy_factor': 1.0, 'csv_stem': '2026_delay_2w'},
    {'id': 'strong', 'label': 'policy_strength_1p2x', 'policy_shift': 0.0, 'policy_factor': 1.2, 'csv_stem': '2026_strength_1p2x'},
    {'id': 'weak', 'label': 'policy_strength_0p8x', 'policy_shift': 0.0, 'policy_factor': 0.8, 'csv_stem': '2026_strength_0p8x'},
]


def _copy_formula_params(base: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    return copy.deepcopy(base if base is not None else FORMULA_PARAMS)


def _build_policy_perturbed_formula(factor: float, months: Sequence[int]) -> Dict[str, Any]:
    fp = _copy_formula_params()
    pe = list(fp.get('policy_effects') or [])
    for m in months:
        idx = int(m) - 1
        if 0 <= idx < len(pe):
            pe[idx] = float(pe[idx]) * float(factor)
    fp['policy_effects'] = pe
    return fp


def build_2026_forecast_scenarios() -> List[Dict[str, Any]]:
    full_season = list(range(1, SIMULATION_MONTHS + 1))
    scenarios: List[Dict[str, Any]] = []
    for spec in FORECAST_2026_SCENARIO_SPECS:
        factor = float(spec['policy_factor'])
        fp = (
            _copy_formula_params()
            if abs(factor - 1.0) < 1e-9
            else _build_policy_perturbed_formula(factor, full_season)
        )
        scenarios.append({
            'id': spec['id'],
            'label': spec['label'],
            'policy_shift': float(spec['policy_shift']),
            'policy_factor': factor,
            'csv_stem': spec['csv_stem'],
            'formula_params': fp,
        })
    return scenarios

SURVEY_GROUP_LABELS = {
    'child': 'Children',
    'adult': 'Adults',
    'elderly': 'Older adults',
}

SURVEY_FACTOR_LABELS = {
    'flu': 'Influenza/ILI',
    'peer': 'Peer influence',
    'pol': 'Policy communication',
    'cog': 'Individual cognition',
}


def print_survey_weight_summary(
    formula_params: Optional[Dict[str, Any]] = None,
    scenario_label: Optional[str] = None,
) -> None:
    if formula_params is None:
        formula_params = FORMULA_PARAMS

    print("\n" + "=" * 72)
    if scenario_label:
        print(f"[{scenario_label}] Vaccination-decision parameters (Logistic formula)")
    else:
        print("Vaccination-decision parameters (Logistic formula)")
    print("Formula: z = b_g + beta*ili_trend + eta*ili_relative + gamma*B_neighbor + policy_m + prior_g*I")
    print("=" * 72)

    print(
        f"Shared coefficients: beta={float(formula_params.get('beta_ili_trend', 0.0)):.3f}, "
        f"eta={float(formula_params.get('eta_ili_relative', 0.0)):.3f}, "
        f"γ={float(formula_params.get('gamma_neighbor', 0.0)):.3f}"
    )
    print("-" * 72)
    print(f"{'Group':<14}{'b_g':>12}{'prior_g':>12}")
    print("-" * 32)
    for grp, label in (('child', 'Children'), ('adult', 'Adults'), ('elderly', 'Older adults')):
        b_g, _, _ = _formula_group_keys(formula_params, grp)
        prior_g = float(formula_params.get(f'prior_{grp}', 0.0))
        print(f"{label:<8}{b_g:>12.3f}{prior_g:>12.3f}")

    pe = list(formula_params.get('policy_effects') or [])
    if pe:
        print("\nMonthly policy parameter policy_m:")
        for i, value in enumerate(pe[:SIMULATION_MONTHS], start=1):
            month_name = SIMULATION_MONTH_LABELS[i - 1] if i - 1 < len(SIMULATION_MONTH_LABELS) else f"Month {i}"
            print(f"  {month_name}: {float(value):+.3f}")
    print("=" * 72)


def print_and_export_factor_weight_summary(
    nodes_template: Dict[int, Dict[str, Any]],
    policy_half_month_shift: float = 0.0,
    scenario_label: str = "baseline forecast",
    output_dir: Optional[str] = None,
    formula_params: Optional[Dict[str, Any]] = None,
) -> None:
    if formula_params is None:
        formula_params = FORMULA_PARAMS

    rng = random.Random(SIMULATION_CONFIG['network_seed'] + 9090)
    nodes = copy.deepcopy(nodes_template)
    for node in nodes.values():
        node['health_status'] = None
    nodes, _ = assign_health_status(nodes, rng=rng)

    factors = ('flu', 'peer', 'pol', 'cog', 'prior')
    factor_labels = {
        'flu': 'Current influenza activity',
        'peer': 'Peer recommendation/influence',
        'pol': 'Policy communication',
        'cog': 'Individual cognition',
        'prior': 'Prior vaccination',
    }
    group_labels = {
        'child': 'Children',
        'adult': 'Adults',
        'elderly': 'Older adults',
    }
    questionnaire_factors = {
        'child': ('flu', 'peer', 'pol'),
        'adult': ('peer', 'pol', 'cog'),
        'elderly': ('peer', 'pol', 'cog'),
    }
    buckets: Dict[str, Dict[str, List[float]]] = {
        group: {factor: [] for factor in factors}
        for group in ('child', 'adult', 'elderly')
    }
    midmonth_steps = {
        (m - 1) * STEPS_PER_MONTH + STEPS_PER_MONTH // 2
        for m in range(1, SIMULATION_MONTHS + 1)
    }

    for step in range(1, DEFAULT_SIMULATION_STEPS + 1):
        month_params = _month_params_for_step(step, MONTHLY_PARAMS, policy_half_month_shift)
        is_daytime = step % 2 == 1
        if is_daytime:
            calculate_payoffs(nodes, is_daytime, month_params, step)

        if step in midmonth_steps and not is_daytime:
            for node_id, node in nodes.items():
                if node.get('health_status') != 'susceptible' or node.get('post_vaccine_susceptible'):
                    continue
                group = _node_population_group(node.get('type', ''))
                _, components = compute_vaccination_logit_components(
                    node_id,
                    nodes,
                    month_params,
                    formula_params,
                )
                for factor in factors:
                    buckets[group][factor].append(float(components.get(factor, 0.0)))

        if not is_daytime:
            simulate_vaccination(
                nodes,
                rng=rng,
                month_params=month_params,
                formula_params=formula_params,
                step=step,
            )
        simulate_infection_and_recovery(nodes, is_daytime, rng=rng, month_params=month_params, step=step)

    def _share_line(share: Dict[str, float]) -> str:
        ordered = sorted(share.items(), key=lambda kv: kv[1], reverse=True)
        return " > ".join(f"{factor_labels[k]}={v * 100:.1f}%" for k, v in ordered)

    lines: List[str] = []
    lines.append("=" * 80)
    lines.append(f"Factor Contribution Weights by Population Group: {scenario_label}")
    lines.append("=" * 80)
    lines.append("Note: weights are normalized by each factor's mean absolute logit contribution; individual cognition is 0 when absent from the current primary formula.")

    for group in ('child', 'adult', 'elderly'):
        mean_abs = {
            factor: float(np.mean(np.abs(values))) if values else 0.0
            for factor, values in buckets[group].items()
        }
        model_denom = sum(mean_abs.values())
        model_share = {
            factor: (mean_abs[factor] / model_denom if model_denom > 1e-12 else 0.0)
            for factor in factors
        }
        q_keys = questionnaire_factors[group]
        q_denom = sum(mean_abs[factor] for factor in q_keys)
        q_share = {
            factor: (mean_abs[factor] / q_denom if q_denom > 1e-12 else 0.0)
            for factor in q_keys
        }

        lines.append("")
        lines.append(f"[{group_labels[group]}]")
        lines.append("  Survey-factor weights (excluding prior vaccination): " + _share_line(q_share))
        lines.append("  Model-term contribution shares (including prior vaccination): " + _share_line(model_share))
        detail = ", ".join(f"{factor_labels[factor]}={mean_abs[factor]:.4f}" for factor in factors)
        lines.append("  Mean absolute logit contributions: " + detail)

    text = "\n".join(lines)
    print("\n" + text)

    target_dir = output_dir if output_dir is not None else OUTPUT_CONFIG.get('output_dir', '')
    filename = "factor_contribution_weights_by_population_group.txt"
    try:
        if target_dir:
            os.makedirs(target_dir, exist_ok=True)
            report_path = os.path.join(target_dir, filename)
        else:
            report_path = filename
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(text)
        print(f"Factor contribution weights by population group saved to: {os.path.abspath(report_path)}")
    except OSError as e:
        print(f"Warning: failed to save factor-weight report: {e}")


def execute_policy_scenario(
    nodes_template,
    initial_summary,
    policy_half_month_shift: float,
    scenario_label: str,
    output_dir: str,
    num_runs: int = 20,
    num_processes: int = 8,
    formula_params: Optional[Dict[str, Any]] = None,
    csv_stem: Optional[str] = None,
    article_csv_dir: Optional[str] = None,
) -> Dict[int, Dict[str, Any]]:
    if formula_params is None:
        formula_params = FORMULA_PARAMS
    if article_csv_dir is None:
        article_csv_dir = OUTPUT_CONFIG.get('article_csv_dir') or ''
    export_plots = bool(OUTPUT_CONFIG.get('export_monthly_plots', True))
    export_excel = bool(OUTPUT_CONFIG.get('export_excel', False))
    OUTPUT_CONFIG['output_dir'] = output_dir
    try:
        os.makedirs(output_dir, exist_ok=True)
    except OSError as e:
        print(f"Warning: could not create output directory '{output_dir}': {e}")

    print("\n" + "#" * 80)
    print(f"[{scenario_label}] policy_half_month_shift = {policy_half_month_shift}")
    pe = list(formula_params.get('policy_effects') or [])
    base_pe = list(FORMULA_PARAMS.get('policy_effects') or [])
    if pe and base_pe and len(pe) == len(base_pe):
        if any(abs(float(a) - float(b)) > 1e-9 for a, b in zip(pe, base_pe)):
            print("Policy parameters: full-season policy_effects were scaled for this scenario")
    print(f"Output directory: {os.path.abspath(output_dir)}")
    print("#" * 80)

    print(f"\nStarting {num_runs} simulations ({scenario_label})")
    print(f"Running in parallel with {num_processes} processes")

    if num_processes > 1:
        args_list = [
            (
                run_index,
                nodes_template,
                SIMULATION_CONFIG['num_steps'],
                MONTHLY_PARAMS,
                SIMULATION_CONFIG['network_seed'] + run_index * 10000,
                policy_half_month_shift,
                formula_params,
            )
            for run_index in range(num_runs)
        ]
        with mp.Pool(processes=num_processes) as pool:
            results = pool.map(run_single_simulation_wrapper, args_list)
        results.sort(key=lambda x: x[0])
        all_time_series_data = [result[1] for result in results]
    else:
        all_time_series_data = []
        for run in range(num_runs):
            print(f"Running simulation {run + 1}/{num_runs}...")
            _, ts = run_single_simulation_wrapper(
                (
                    run,
                    nodes_template,
                    SIMULATION_CONFIG['num_steps'],
                    MONTHLY_PARAMS,
                    SIMULATION_CONFIG['network_seed'] + run * 10000,
                    policy_half_month_shift,
                    formula_params,
                )
            )
            all_time_series_data.append(ts)

    print(f"\n{scenario_label}: {num_runs} simulations completed")

    all_steps = set()
    for data in all_time_series_data:
        all_steps.update(d['step'] for d in data)
    steps = sorted(all_steps)

    avg_time_series_data = []
    for step in steps:
        infected_values = []
        vaccinated_values = []
        recovered_values = []
        susceptible_values = []
        for time_series_data in all_time_series_data:
            data_by_step = {d['step']: d for d in time_series_data}
            if step in data_by_step:
                infected_values.append(data_by_step[step]['infected'])
                vaccinated_values.append(data_by_step[step]['vaccinated'])
                recovered_values.append(data_by_step[step]['recovered'])
                susceptible_values.append(data_by_step[step]['susceptible'])
        if infected_values:
            avg_time_series_data.append({
                'step': step,
                'susceptible': np.mean(susceptible_values),
                'infected': np.mean(infected_values),
                'vaccinated': np.mean(vaccinated_values),
                'recovered': np.mean(recovered_values),
            })

    monthly_increments = calculate_monthly_increments(
        all_time_series_data,
        num_runs=num_runs,
        initial_summary=initial_summary,
    )
    print_survey_weight_summary(formula_params=formula_params, scenario_label=scenario_label)
    print_monthly_increments(monthly_increments)
    if export_plots:
        plot_monthly_vaccination_ci_by_group(
            monthly_increments,
            num_runs=num_runs,
            real_values={},
        )
    export_monthly_vaccination_ci_csv(
        monthly_increments,
        scenario_label=scenario_label,
        output_dir=output_dir,
    )
    if article_csv_dir:
        export_monthly_vaccination_ci_csv(
            monthly_increments,
            scenario_label=scenario_label,
            output_dir=article_csv_dir,
            csv_stem=csv_stem or scenario_label,
        )
    print_and_export_factor_weight_summary(
        nodes_template,
        policy_half_month_shift=policy_half_month_shift,
        scenario_label=scenario_label,
        output_dir=output_dir,
        formula_params=formula_params,
    )
    if export_excel:
        export_to_excel(avg_time_series_data)
    return monthly_increments

if __name__ == "__main__":
    mp.freeze_support()

    network_rng = random.Random(SIMULATION_CONFIG['network_seed'])

    family_dist = calculate_family_distribution()
    print_family_distribution(family_dist)

    nodes, child_indices, _, _, _ = create_node_network(family_dist, rng=network_rng)
    nodes = add_school_connections(nodes, child_indices, rng=network_rng)
    nodes = add_company_connections(nodes, rng=network_rng)
    nodes = add_elderly_connections(nodes, rng=network_rng)
    assign_prior_vaccination_history(nodes, rng=network_rng)
    assign_cognition_traits(nodes, rng=network_rng)
    nodes, initial_summary = assign_health_status(nodes, rng=network_rng)

    print_initial_health_summary(initial_summary)
    print_prior_vaccination_summary(nodes)
    print_prior_vaccination_ratio_diagnostic(FORMULA_PARAMS)

    load_weekly_ili_caches()

    print("\n" + "=" * 60)
    print("Monthly Parameter Configuration")
    print("=" * 60)
    print("Vaccination rule: Logistic formula (FORMULA_PARAMS)")
    policy_scenarios = build_2026_forecast_scenarios()
    article_csv_dir = OUTPUT_CONFIG.get('article_csv_dir') or ''
    print(
        "Policy scenarios will run in sequence: "
        + ", ".join(sc['label'] for sc in policy_scenarios)
        + f" ({len(policy_scenarios)} total)."
    )
    if article_csv_dir:
        print(f"Article CSV output directory: {os.path.abspath(article_csv_dir)}")
    print(f"Simulation-results directory: {os.path.abspath(OUTPUT_CONFIG['output_dir'])}")
    print(
        "ILI mapping: step -> day -> simulation week (week_index_from_step) -> CSV epidemiological week. "
        "Simulation week 1 maps to table week 31, followed by 32...52 -> 1...13 "
        "(see csv_epi_week_from_simulation_week / SIMULATION_EPI_WEEK_SEQUENCE)."
    )
    wcsv = ILI_CONFIG.get('weekly_csv') or '(not configured)'
    print(f"  weekly_csv: {wcsv}")
    print("Autonomous infection rate r_auto is calculated from the current week's lambda (CSV); r_auto=0 for a step if CSV loading fails or the week is missing.")
    print(f"  ILI CSV load failed: {ILI_WEEKLY_CSV_LOAD_FAILED}")
    print("Note: P_n (neighbor transmission) is printed by month below; r_auto is not configured monthly and is derived per step from the ILI CSV.")
    for month in range(1, SIMULATION_MONTHS + 1):
        params = MONTHLY_PARAMS[month]
        print(f"\nMonth {month} (time steps {MONTH_BOUNDARIES[month - 1][0]}-{MONTH_BOUNDARIES[month - 1][1]}):")
        pn = params.get('p_single_infection', INFECTION_CONFIG.get('p_single_infection', 0.05))
        print(f"  P_n (single-infected-neighbor transmission probability): {pn:.6f}")

    num_runs = 20
    num_processes = 8

    nodes_template = copy.deepcopy(nodes)
    for node_id in nodes_template:
        nodes_template[node_id]['health_status'] = None

    base_output_dir = OUTPUT_CONFIG['output_dir']
    for scenario in policy_scenarios:
        scenario_label = scenario['label']
        scenario_dir = (
            os.path.join(base_output_dir, scenario_label)
            if base_output_dir
            else scenario_label
        )
        execute_policy_scenario(
            nodes_template,
            initial_summary,
            policy_half_month_shift=float(scenario['policy_shift']),
            scenario_label=scenario_label,
            output_dir=scenario_dir,
            num_runs=num_runs,
            num_processes=num_processes,
            formula_params=scenario['formula_params'],
            csv_stem=scenario['csv_stem'],
            article_csv_dir=article_csv_dir,
        )

    print("\n" + "=" * 80)
    print("All 2026 policy-scenario simulations completed.")
    if article_csv_dir:
        print(f"Article CSV files: {os.path.abspath(article_csv_dir)}")
        for sc in policy_scenarios:
            print(f"  - {sc['csv_stem']}.csv")
    print("Simulation subdirectories:")
    for scenario in policy_scenarios:
        sub = (
            os.path.join(base_output_dir, scenario['label'])
            if base_output_dir
            else scenario['label']
        )
        print(f"  - {scenario['label']}: {os.path.abspath(sub)}")
    print("=" * 80)
